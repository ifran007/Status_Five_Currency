import openpyxl
from openpyxl.styles import Alignment


def getRowCount(file, sheetName):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetName]
    return sheet.max_row


def getColCount(file, sheetName):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetName]
    return sheet.max_column


def readData(file, sheetName, rowno, colno):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetName]
    return sheet.cell(rowno, colno).value


def writeData(file, sheetName, rowno, colno, data):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetName]
    sheet.cell(rowno, colno).value = data
    workbook.save(file)


def readsinglerow(file, sheetName, rownum, colnum):
    single_row = []
    for i in range(1, colnum + 1):
        data = readData(file, sheetName, rownum, i)
        single_row.append(data)
    return single_row


def readsinglecol(file, sheetName, srow, maxrownum, colnum):
    single_col = []
    for i in range(srow, maxrownum + 1):
        data = readData(file, sheetName, i, colnum)
        single_col.append(data)
    return single_col


def writelistoflist(file, sheetName, data):
    wb = openpyxl.load_workbook(file)
    sheet = wb[sheetName]
    row = 1
    for element in data:
        col = 1
        for d in element:
            if d:
                sheet.cell(row, col).value = d
            col += 1
        row += 1
    wb.save(file)


def writesinglerow(file, sheetName, rownum, colnum, scol, data):
    wb = openpyxl.load_workbook(file)
    sheet = wb[sheetName]
    for i in range(1, colnum + 1):

        sheet.cell(rownum, i + scol).value = data[i - 1]
        # print(data[i - 1])
    wb.save(file)


def writesinglecol(file, sheetName, rownum, colnum, srow, data):
    wb = openpyxl.load_workbook(file)
    sheet = wb[sheetName]
    for i in range(1, rownum + 1):
        sheet.cell(i + srow, colnum).value = str(data[i - 1])
        # print(data[i - 1])
    wb.save(file)


def mergecell(file, sheetName, mcell1, mcell2, rownum, colnum, ):
    wb = openpyxl.load_workbook(file)
    sheet = wb[sheetName]
    sheet.merge_cells(f'{mcell1}:{mcell2}')
    cell = sheet.cell(rownum, colnum)
    cell.alignment = Alignment(horizontal='center', vertical='center')
    wb.save(file)


def writeinmergecell(file, sheetName, mcell1, mcell2, rownum, colnum, scol, data):
    mergecell(file, sheetName, mcell1, mcell2, rownum, scol + 1)
    writesinglerow(file, sheetName, rownum, colnum, scol, data)


def createlktestresultExcelfile(filepath):
    wb = openpyxl.Workbook()
    # ws = wb.active
    wb.create_sheet("GetAppointments", 0)
    wb.create_sheet("GetMedications", 1)
    wb.create_sheet("GetAllergies", 2)
    wb.create_sheet("GetImmunizations", 3)
    wb.create_sheet("GetLabResults", 4)
    sheet = wb['Sheet']
    wb.remove(sheet)
    wb.save(filename=filepath)


def remove_items(list1, item):
    res = [i for i in list1 if i != item]
    return res
